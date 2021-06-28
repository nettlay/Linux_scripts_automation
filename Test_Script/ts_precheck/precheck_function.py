#!/usr/bin/python3 env
# Automated test script required functions

import pyautogui
from time import sleep
import datetime
import subprocess
import os, sys
import openpyxl
import cv2,re
import yaml
from Common.common_function import argv_filter
from Common import common_function

log = common_function.log


def get_current_dir():
    if os.path.dirname(sys.argv[0]) == '':
        return os.getcwd()
    else:
        return os.path.dirname(sys.argv[0])


# open a window
def open_window(window_name):
    log.info("press ctrl alt s to open window {}".format(window_name))
    pyautogui.hotkey("ctrl", "alt", "s")
    sleep(1)
    pyautogui.typewrite(window_name, interval=0.1)
    sleep(0.5)
    pyautogui.press("enter")
    sleep(3)
    log.info("open window {} end".format(window_name))


def now():
    now_time = datetime.datetime.today().strftime('%Y-%m-%d_%H-%M-%S')
    return now_time


# search picture on now screen. default search 3 times.
def check_window(picture, times=1):
    now_resolution = str(pyautogui.size()[0]) + 'x' + str(pyautogui.size()[1])
    if os.path.exists(os.path.join(get_current_dir(), r'Test_Data/td_precheck', now_resolution)):
        p = os.path.join(get_current_dir(), r'Test_Data/td_precheck', now_resolution, 'picture', picture)
    else:
        p = os.path.join(get_current_dir(), r'Test_Data/td_precheck', '1920x1080', 'picture', picture)
    # p = os.path.join(get_current_dir(), 'Test_Data', now_resolution, 'picture', picture)
    pyautogui.screenshot("test.png")
    sleep(0.1)
    samp = cv2.imread("test.png")
    temp = cv2.imread(p)
    sleep(0.1)
    os.popen("rm test.png")
    while times:
        res = cv2.matchTemplate(samp, temp, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        sleep(0.1)
        temp_x = temp.shape[0]
        temp_y = temp.shape[1]
        if max_val > 0.95:
            sleep(0.1)
            x = max_loc[0] + int(temp_x / 2)
            y = max_loc[1] + int(temp_y / 2)
            return x, y
        else:
            times = times - 1
            if times == 0:
                return None


# get mouse position in the click_position.yaml
def mouse_position():
    resolution = str(pyautogui.size()[0]) + 'x' + str(pyautogui.size()[1])
    if os.path.exists(os.path.join(get_current_dir(), r'Test_Data\td_precheck', resolution)):
        yaml_file = os.path.join(get_current_dir(), r'Test_Data\td_precheck', resolution, 'yaml_file', 'click_position.yaml')
    else:
        yaml_file = os.path.join(get_current_dir(), r'Test_Data\td_precheck', '1920x1080', 'yaml_file', 'click_position.yaml')
    file = open(yaml_file, 'r')
    p = yaml.full_load(file)
    return p


def primary_card_label():
    output = os.popen('fdisk -l | grep /dev/ | grep -v Disk').readlines()
    line_boot = ''
    for line in output:
        # if '*' in line:
        if 'Linux' in line:
            line_boot = line
            break
    if line_boot != '':
        label_boot = line_boot.split(' ')[0]
        if 'mmcblk' in label_boot or 'nvme' in label_boot:
            pri_card_label = label_boot[:-2]
        else:
            pri_card_label = label_boot[:-1]
    else:
        pri_card_label = ''
    return pri_card_label


# check system is thinpro or smart zreo
def thinpro_or_smartzero():
    os_configuration = subprocess.getoutput("mclient --quiet get tmp/SystemInfo/general/ProductConfig")
    if os_configuration == "zero":
        return "smart zero"
    elif os_configuration == "standard":
        return "thinpro"
    else:
        log.error("can't tell if the system is thinpro or smart zero")
        return False


def get_ip():
    wired_status = subprocess.getoutput("mclient --quiet get tmp/NetMgr/eth0/IPv4/status")
    if wired_status == "1":
        sys_eth0_ip = subprocess.getoutput("ifconfig | grep eth0 -A 1 | grep -i 'inet addr'")
        eth0_ip = sys_eth0_ip.strip().split()[1].split(":")[1]
        return eth0_ip

    wireless_status = subprocess.getoutput("mclient --quiet get tmp/NetMgr/wlan0/IPv4/status")
    if wireless_status == "1":
        sys_wlan0_ip = subprocess.getoutput("ifconfig | grep wlan0 -A 1 | grep -i 'inet addr'")
        wlan0_ip = sys_wlan0_ip.strip().split()[1].split(":")[1]
        return wlan0_ip

    else:
        with os.popen("ifconfig") as f:
            string = f.read()
            eth0_ip = re.findall("eth0.*?addr.*?(\d+\.\d+\.\d+\.\d+).*?lo", string, re.S)
            if eth0_ip:
                eth0_ip = eth0_ip[0]
            else:
                eth0_ip = get_ip_from_yml()
        return eth0_ip


def get_ip_from_yml():
    report_path = os.path.join(get_current_dir(), 'Test_Report')
    files = os.walk(report_path)
    for file in files:
        if '.yaml' in file[2][0]:
            ip_addr = file[2][0].split('.')[:-1]
            return '.'.join(ip_addr)
    return 'null'


class SwitchThinProMode:
    '''
    switch thinpro to admin mode or user mode
    SwitchThinProMode(switch_to='admin')
    '''
    def __init__(self, switch_to):

        self.switch_to = switch_to
        self.switch_mode()

    @staticmethod
    def current_mode():
        s = os.path.exists('/var/run/hptc-admin')
        if s:
            return 'admin'
        else:
            return 'user'

    @staticmethod
    def judge_has_root_pw():
        s = subprocess.call("hptc-passwd-default --root --check >/dev/null 2>&1", shell=True)
        if s:
            return True
        else:
            return False

    def switch_mode(self):
        if self.switch_to == 'user':
            if SwitchThinProMode.current_mode() == 'user':
                log.info("now is user mode")
                return True
            if SwitchThinProMode.current_mode() == 'admin':
                os.popen('hptc-switch-admin')
                sleep(2)
                if SwitchThinProMode.current_mode() == 'user':
                    # print('switch to user mode success')
                    log.info("switch to user mode success")
                    return True
                else:
                    log.error("switch to user mode fail")
                    return False

        if self.switch_to == 'admin':
            if SwitchThinProMode.current_mode() == 'admin':
                log.info("now is admin mode")
                return True
            if SwitchThinProMode.current_mode() == 'user':
                if SwitchThinProMode.judge_has_root_pw():
                    os.popen('hptc-switch-admin')
                    sleep(2)
                    pyautogui.typewrite('1')
                    sleep(1)
                    pyautogui.hotkey('enter')
                    sleep(2)
                    if SwitchThinProMode.current_mode() == 'admin':
                        log.info("switch to admin mode success")
                        return True
                    else:
                        log.error("switch to admin mode fail")
                        return False

                if not SwitchThinProMode.judge_has_root_pw():
                    os.popen('hptc-switch-admin')
                    sleep(2)
                    pyautogui.typewrite('1')
                    sleep(1)
                    pyautogui.hotkey('tab')
                    hash(1)
                    pyautogui.typewrite('1')
                    sleep(1)
                    pyautogui.hotkey('enter')
                    sleep(2)
                    if SwitchThinProMode.current_mode() == 'admin':
                        log.info("switch to admin mode success")
                        return True
                    else:
                        log.error("switch to admin mode fail")
                        return False


def card_size(card_label):
    str_command = 'fdisk -l | grep \'Disk {}\''.format(card_label)
    output = subprocess.getoutput(str_command)
    if not output:
        log.info('No result for querying card label {}'. format(card_label))
        return False
    else:
        size_GiB = output.split(',')[0].split(' ')[-2]
        size_bytes = output.split(',')[1].strip().split(' ')[0]
        # size_GB = size_bytes.split(' ')[0][:-9]
        return size_GiB, size_bytes


def partition_size(card_label):
    str_command = 'fdisk -l | grep {}'.format(card_label)
    output = os.popen(str_command).readlines()
    line_partition = ''
    for line in output:
        if 'Linux' in line:
            line_partition = line
            break
    if line_partition != '':
        part_size_G = line_partition.split(' ')[-3][:-1]
    else:
        part_size_G = ''
    return part_size_G


# Read primary and secondary storage information from DASH matrix excel
def storage_info(platform, config):
    excel = openpyxl.load_workbook(
        r'{}/Test_Data/User_Defined_Data/td_precheck/ThinPro7.1_DASH_Test Matrix.xlsx'.format(get_current_dir()))
    sheet = excel.active
    rows = sheet.max_row
    cols = sheet.max_column
    row_second_storage = 0
    for row in range(1, rows + 1):
        # get secondary storage row
        for col in range(1, cols + 1):
            if sheet.cell(row, col).value == 'Secondary Storage':
                row_second_storage = row
                break
            else:
                continue
        if row_second_storage != 0:
            break
    col_config = 0
    for col in range(1, cols):
        #  get row of specific config and platform
        cell_value = sheet.cell(1, col).value
        if not cell_value:
            continue
        if config.upper() in cell_value.upper() and platform.upper() in cell_value.upper():
            col_config = col
            break
    primary_storage = sheet.cell(row_second_storage - 1, col_config).value
    second_storage = sheet.cell(row_second_storage, col_config).value
    return primary_storage, second_storage

